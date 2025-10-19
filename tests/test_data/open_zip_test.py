from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def zipfile_for_image():
    # 创建一个新的工作簿
    wb = Workbook()
    # 获取活动的工作表
    ws = wb.active

    # 打开TXT文件
    with open(r"D:\SPR\文件\Ab0001 4E-08M_1780866861018255360 20240418155201.txt", 'r') as f:
        # 逐行读取
        for line in f:
            # 根据分隔符分割行
            cells = line.strip().split( )  # 假设使用逗号作为分隔符
            # 将分割后的数据作为一行写入工作表
            ws.append(cells)

    # 自定义单元格样式（可选）
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center')

    # 应用样式到第一行
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment

    # 保存工作簿
    wb.save(r"C:\Users\86155\Desktop\qweqwe.xlsx")
    return r"C:\Users\86155\Desktop\qweqwe.xlsx"