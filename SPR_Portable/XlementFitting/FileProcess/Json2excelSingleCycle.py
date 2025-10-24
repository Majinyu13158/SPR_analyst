import json
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

__all__ = [
    "is_json_structure_right",
    "is_any_OriginalDataList_empty",
    "convert_json_to_excel_single_cycle"
    ]

def is_json_structure_right(json_path):
    try:
        with open(json_path, 'r', encoding='utf-8-sig') as file:
            data = json.load(file)
        
        if not isinstance(data, list):
            print(f"{json_path}无法读取")
            return False
        
        for item in data:
            if not isinstance(item, dict):
                print(f"{json_path}格式不对")
                return False
            
            required_keys = {
                "ExperimentID", "SampleID", "Molecular", "SampleName",
                "Concentration", "ConcentrationUnit", "CombineStartIndex",
                "CombineEndIndex", "OriginalDataList"
            }
            
            if not all(key in item for key in required_keys):
                print(f"{json_path}格式不对")
                return False
            
            if not isinstance(item["OriginalDataList"], list):
                print(f"{json_path}格式不对")
                return False
            
            for original_data in item["OriginalDataList"]:
                if not isinstance(original_data, dict):
                    print(f"{json_path}无法读取")
                    return False
                if not all(key in original_data for key in ["ID", "Test", "Refe"]):
                    print(f"{json_path}格式不对")
                    return False
        
        return True
    
    except (json.JSONDecodeError):
        print(f"{json_path}格式不对")
        raise
    except (FileNotFoundError):
        print(f"{json_path}无法读取")
        raise

def is_any_OriginalDataList_empty(json_path):
    try:
        with open(json_path, 'r', encoding='utf-8-sig') as file:
            data = json.load(file)
        
        if not isinstance(data, list):
            print(f"{json_path}无法读取")
            return False
        
        for item in data:
            if isinstance(item, dict) and "OriginalDataList" in item:
                if item["OriginalDataList"] == []:
                    print(f"{json_path}中的实验{item["ExperimentID"]}存在空的OriginalDataList")
                    return False
        
        return True
    
    except (json.JSONDecodeError, FileNotFoundError):
        raise

def convert_json_to_excel_single_cycle(path_json):
    if not is_json_structure_right(path_json):
        return path_json
    if not is_any_OriginalDataList_empty(path_json):
        return path_json
    path_json = Path(path_json)
    # 读取 JSON 文件
    with open(path_json, 'r', encoding='utf-8-sig') as file:
        data = json.load(file)

    # 创建一个 Excel 工作簿
    wb = Workbook()

    # 删除默认创建的 sheet
    wb.remove(wb.active)

    # 用于存储 CombineStartIndex 和 CombineEndIndex
    combine_indices = []

    # 用于存储信号数据
    signal_data = []

    # 处理每个部分
    for index, item in enumerate(data, start=1):
        
        # 计算 Concentration / Molecular
        conc_mol_ratio = item['Concentration'] / (item['Molecular']*1e+3)
        
        # 创建 DataFrame 来存储 OriginalDataList 数据
        df = pd.DataFrame(item['OriginalDataList'])
        df['Difference'] = df['Test'] - df['Refe']
        df = df.reset_index()  # 添加序号列
        df = df[['index', 'Difference']]  # 只保留序号和差值列
        df = df.rename(columns={'index': 'Rank'})  # 将 'index' 重命名为 'Rank'
        df['Rank'] += 1  # 使 Rank 从 1 开始
        
        
        # 收集 CombineStartIndex 和 CombineEndIndex
        combine_indices.append({
            '浓度': conc_mol_ratio,
            '结合起点': item['CombineStartIndex'],
            '解离起点': item['CombineEndIndex']
        })
        
        # 收集信号数据
        signal_data.append({
            'conc_mol_ratio': conc_mol_ratio,
            'data': df
        })
        
    # 创建信号 sheet
    signal_sheet = wb.create_sheet(title="信号")
    signal_sheet['A1'] = 'time'
    signal_sheet['B1'] = 'signal'

    # 按 conc_mol_ratio 排序信号数据
    signal_data.sort(key=lambda x: x['conc_mol_ratio'])

    # 填充信号数据
    time_offset = 0
    all_differences = []
    all_times = []

    for item in signal_data:
        df = item['data']
        df['time'] = df['Rank'] + time_offset
        all_differences.extend(df['Difference'].tolist())
        all_times.extend(df['time'].tolist())
        time_offset += len(df)
        
    # 创建完整的 DataFrame
    full_df = pd.DataFrame({
        'time': all_times,
        'signal': all_differences
    })

    # 对 signal 列进行归零处理
    full_df['normalized_signal'] = full_df['signal'] - full_df['signal'].iloc[0]

    # 将数据写入信号 sheet
    for _, row in full_df.iterrows():
        signal_sheet.append([row['time'], row['normalized_signal']])

    # 创建 CombineIndices sheet
    combine_sheet = wb.create_sheet(title="测试信息")
    combine_df = pd.DataFrame(combine_indices)

    for r in dataframe_to_rows(combine_df, index=False, header=True):
        combine_sheet.append(r)

    # 保存 Excel 文件
    stem_name = path_json.stem
    parent_dir = path_json.parent
    excel_name = parent_dir / f"{stem_name}.xlsx"
    wb.save(excel_name)

    # print(f"已经保存excel{excel_name}")
    return excel_name

if __name__ == "__main__":
    convert_json_to_excel_single_cycle("ScriptData/sss.json")